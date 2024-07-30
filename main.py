import numpy as np
import matplotlib.pyplot as plt
from tifffile import imread
import openpyxl
import pywt 
from scipy.interpolate import interp1d
from scipy.ndimage import gaussian_filter1d,gaussian_filter
from skimage.feature import peak_local_max
from scipy.signal import find_peaks
from sklearn.metrics import accuracy_score
import glob
import os
import seaborn as sns
import yaml
from tqdm import tqdm
sns.set()
'''
===============================================================
        Arrhythpy, WRITTEN BY Karim Ajmail, 2023
===============================================================
'''

'''
Arrhythpy is a python-based program, that quantifies arrhythmias in oscillatory time series data. Input are confocal line scans, which are averaged along the scanning direction.
It is based on extracting time resolved frequency information by performing the Wavelet transformation.
The resulting coefficient map obtained by such an Wavelet transformation is analysed in the following way:
 1. Local Minima and Maxima are determined with the peak_local_max function (wvl_max / wvl_min)
 2. Since we need pairs of maxima and minima to get one peak, we obtained corresponding maxima  based on the minima.
    This is done by estimating the approximate location of the maxima by interpolation and bisection between two minima (approx_max). 
    The closest local maxmium to this middle point between the two minima (exact_maxima) is the used for further analysis.
 3. In case there are multiple and different maxima between two minima, we have to decide which maximum we choose. 
    This criterium is simple: Choose the maximum with smallest distance to the adjacent larger minmum. Choosing the minimum,that is right to the maxima, makes sense, since one peak consists of a maximum first and then a minimum.
    In the case of the maxima, we choose the distance to the minima as the criterium, since we want adjacent maxima and minima and a lot of maxima are low frequecy components of double/triple  peaks.
 4. Of course there can also be multiple minima between two maxima. In this case we choose the lowest i.e minimum with highest amplitude.
    For minima, low frequency information is mostly due to a lack of a peak, hence it makes sence to choose the lowest minimum.
 5. Each pair of maximum and minimum is averaged and interpolated (nearest). Note, that due to the nature of the mexican hat wavelet the minimum value is counted double in the average.
 6. The obtained width information is converted into frequency information.

    Side Note: the ending _x denotes the time coordinate in the coefficient map (wvl)

The obtained time resolved frequency information is compared to the given experimatal pacing frequency.
The positve deviation (higher frequency) is averaged and saved as the tachycardial contribution to the arrhyhtmia.
Similarly, for bradicardial contribution the negative deviation is used. Tachycardic and bradycardic deviation from pacing frequency are used to clasify into Bradycardia and Tachycardia respectivly.

To quantify how periodic i.e. self-similar the signal the autocrrelation with periodic boundary condiction is used. The periodic BC is implemented by cropping the signal at the first and last root.
Both from negative to positive and from positive to negative zero crossings are tested and the maximal correlation is used. This is basically a fail safe in case one of those fails. The scale of this parameters
is inverted to get a parameters ranging from 0 to 1, where 1 is the most arrhythmic ones.

For a measure of arrhythmia in general independent of the pacing frequency we use a combination of two parameters namely the Inverse Mean Correlation (IMC) and the Variance of Frequency (VF).
These two are simple averaged with a given weight (usually set to 0.5). The VF is also cropped to a range of [0,1]. We did not scale the VF by its variance to avoid that the metric changes for each dataset.
Based on this metric, each transient is classified into arrhythmic and rythmic independent of pacing frequency. 
Therefore, there are 6 classes: Tachycardic rhythmic, Tachycardic arrhythmic, rhytmic, arrythmic, bradycardic rhythmic, bradycardic arrhythmic

All these analysed parameters together with the appropriate plots are saved in an excel sheet.
'''

def run(files,frequency,threshold_freq = .1, threshold_arrythmia = .1, duration = 10.,sigma0 = .05,prominence=0.0005,threshold_abs = .1,varfreq_weight = 0.5,scale = 1.,path = None,show = False, save_mode = False):
    '''
    Parameter:
    - files:                list of files to be analyses. Must be line scan images (either tif or lsm files)
    - frequency:   	        float; pacing frequency or known eigenfrequency of cells
    - threshold_freq:       float; threshold for classification into tachy- & bradycardic based on frequency deviation from pacing frequency
    - threshold_arrythmia:  float; threshold for classification inot rhythmic and arrythmic based on the combined arrythmia measure form IMC and VF
    - duration:             float; duration of the transient in signal (needed for unit conversion)
    - sigma0:  	            float; Standard deviation of the Gaussian used for the gaussian filter. Too low values will increase the noise in the signal,while too large values will eliminate highly tachycardic signals. 
                                sigma is changed depending on the largest amplitude of the Fourier specturm to adapt it to the acutal frequency in the signal.
    - threshold_abs:        float; threshold for local extrema i.e. minimum height of the peak
    - varfreq_weight:       float; weight in weithed average of VF and IMC
    - scale:                float; the VF is scaled by this number to have IMC and VF roughly on th same scale.
    - show:                 bool; if True, the results for each transient are displayed.
    '''

    classes = {'Rhythmic' : 0, 'Tachycardic Rhythmic' : 1, 'Bradycardic Rhythmic' : 2, 'Arrhythmic' : 3, 'Tachycardic Arrhythmic': 4, 'Bradycardic Arrhythmic' : 5}

    if not path:
        path = '\\'.join(files[0].split('\\')[:-1])

    if len(files) > 0:
        position = 1                        # initialise position in excel spread sheet
        workbook = openpyxl.Workbook()      # open excel spread sheet
        sheet = workbook.active             # and activate it

        histogram = np.full([len(files),2],-1.)
        distribution = np.zeros(6)
        weights = np.full(len(files),np.nan)
        sigmas = np.full(len(files),np.nan)
        reds = np.full(len(files),np.nan)
        croppings = np.full(len(files),np.nan,dtype="S10")
        
        path_figures = os.path.join(path,'figures')
        c = -1                               # counter, initilased at -1 so that it starts at 0
        for num in tqdm(range(len(files)),desc='Arrhythpy: '):
            file = files[num]
            #print('File: '+os.path.basename(file) +f' ({num+1} of {len(files)})')
            try:
                data_type = file.split('.')[-1]
                if data_type =='lsm' or data_type == 'tif':
                    img = imread(file)[:,0,:]
                    intensity = np.mean(img,axis=1)
                    c += 1
                elif data_type == 'csv':
                    intensity = np.genfromtxt(file, delimiter=' ')
                    c += 1
                else:
                    print(f'\tWarning: Data type \'.{data_type}\' not supported!')
                    continue

                red = int(len(intensity) // 5000)                                                    # this is to reduce computation time
                if red < 1:
                    red = 1
            
                intensity = intensity[::red]
                time = np.linspace(0,duration,len(intensity))
 
                fft = abs(np.fft.rfft(intensity-intensity.mean())) 
                freq_fft = np.fft.fftfreq(len(intensity), d=duration/len(intensity))[np.argmax(fft)] / 2        # get maximum amplitude of fourier spectrum , /2 since rfft crops the symetric spectrum at half

                if frequency is None:
                    frequency = freq_fft
                    print('Warning: No pacing/eigen frequency given. Tachycardia and bradycardia are calculated with highest peak in fourier spectrum. \n')

                width_used = len(intensity)/(duration*frequency) / 4                                        # for mexical hat wavelet width is defined like this
                #end_width = max(width_used+50,len(intensity)/(duration*freq_fft)/3)  # largest width used for wavelet tranform (depending on ratio between given frequency and frequency of main fourier coefficient)
                end_width = len(intensity)/(duration*freq_fft) / 3  # largest width used for wavelet tranform (depending on ratio between given frequency and frequency of main fourier coefficient)

                sigma = int (len(intensity)/(duration*freq_fft) *sigma0)             # width of gaussian is adapted to main frequency component
                filtered = gaussian_filter1d(intensity,sigma=sigma,mode='nearest')   # gaussian filter to smooth

                scales = np.linspace(1,end_width,500)
                wvl,_ = pywt.cwt(intensity-np.mean(intensity),scales,'mexh',sampling_period=duration/len(intensity))  # continious wavelet transform
                wvl /= (wvl.max() - wvl.min())     
                wvl = gaussian_filter(wvl,10)     

                # 1.
                mins = peak_local_max(-wvl,min_distance=10,threshold_abs=threshold_abs,exclude_border=False)    # locate local minima
                wvl_min_x,wvl_min = mins[:,1],scales[mins[:,0]] # cut off the sides to avoid border artifacts
                wvl_min = wvl_min[np.argsort(wvl_min_x)]                                                        # sort along time
                wvl_min_x = wvl_min_x[np.argsort(wvl_min_x)]
                wvl_min = wvl_min[wvl_min_x<len(intensity)-width_used*2]
                wvl_min_x = wvl_min_x[wvl_min_x<len(intensity)-width_used*2]

                maxs = peak_local_max(wvl,min_distance=10,threshold_abs=threshold_abs ,exclude_border=False)    # locate local maxima
                wvl_max_x,wvl_max = maxs[:,1],scales[maxs[:,0]] # cut off the sides to avoid border artifacts
                wvl_max = wvl_max[np.argsort(wvl_max_x)]                                                        # sort along time
                wvl_max_x = wvl_max_x[np.argsort(wvl_max_x)]
                wvl_max = wvl_max[wvl_max_x>width_used*2]
                wvl_max_x = wvl_max_x[wvl_max_x>width_used*2]

                # 2.
                if len(wvl_min) > 1: 
                    f_diff = interp1d(wvl_min_x,wvl_min)                                                        # linear interpolation
                    wvl_min = wvl_min[wvl_min_x>= wvl_max_x[0]]                                                 # delete minima left to first maximum
                    wvl_min_x = wvl_min_x[wvl_min_x>= wvl_max_x[0]]                                             # since we have to start with a maximum
                    exact_maxima = np.zeros(wvl_min_x.shape)
                    exact_maxima_x = np.zeros(wvl_min_x.shape)

                    for i,_ in enumerate(wvl_min_x):
                        if i == 0:                                                                              # for the first minimum we don't have a minimum before that -> just use the average distance
                            approx_max_x = wvl_min_x[i] - .5*np.mean(np.diff(wvl_min_x))
                            approx_max = wvl_min[i]
                        else:
                            approx_max_x = wvl_min_x[i] - .5 * np.diff([wvl_min_x[i-1],wvl_min_x[i]])[0]        # get middle point between two minima for obtain an approximate maximum
                            approx_max = f_diff(approx_max_x)  

                        dist = (wvl_max_x[wvl_max_x < wvl_min_x[i]] - approx_max_x)**2 + (wvl_max[wvl_max_x < wvl_min_x[i]] - approx_max)**2 # get dsitance betweeen the approximate maxima and the actual maxima
                        exact_maxima[i] = wvl_max[np.argmin(dist)]                                              # choose the maximum closest to the approximate one
                        exact_maxima_x[i] = wvl_max_x[np.argmin(dist)]
                else:                                                                                           # if only one entery, then I don't need this and I just take the global maximum
                    exact_maxima,exact_maxima_x = np.where(wvl==wvl.max())
                
                exact_maxima_x,idx_unique = np.unique(exact_maxima_x,return_index=True)                         # reduce the array to only contain unique enteries
                exact_maxima = exact_maxima[idx_unique]

                # 3.
                for i in range(len(wvl_min)-1):
                    cut = exact_maxima[(exact_maxima_x > wvl_min_x[i]) & (exact_maxima_x < wvl_min_x[i+1])]     # cut between two minima 
                    cut_x = exact_maxima_x[(exact_maxima_x > wvl_min_x[i]) & (exact_maxima_x < wvl_min_x[i+1])] # to see if there is more than one minimum
                    if len(cut) >1:                                                                             # if so 
                        d = (cut_x - wvl_min_x[i+1])**2 + (cut - wvl_min[i+1])**2                               # calculate the distance                     
                        exact_maxima = np.delete(exact_maxima,(exact_maxima_x[:, None] == cut_x[cut_x != cut_x[np.argmin(d)]]).argmax(axis=0)) # delete those maxima, that are not the closest one
                        exact_maxima_x = np.delete(exact_maxima_x,(exact_maxima_x[:, None] == cut_x[cut_x != cut_x[np.argmin(d)]]).argmax(axis=0))

                # 4.
                new_wvl_min = np.zeros(exact_maxima.shape)
                new_wvl_min_x = np.zeros(exact_maxima.shape)
                if len(exact_maxima) > 1:
                    for i in range(len(exact_maxima)-1):

                        cut = (scales[:, None] == wvl_min[(wvl_min_x > exact_maxima_x[i]) & (wvl_min_x < exact_maxima_x[i+1])]).argmax(axis=0) # float scale values are converted into indeces
                        cut_x = wvl_min_x[(wvl_min_x > exact_maxima_x[i]) & (wvl_min_x < exact_maxima_x[i+1])]          # to see if there ise more than one minimum between two maxima
                        ampl = wvl[cut,cut_x]                                                                           # if so
                        new_wvl_min[i] = scales[cut[np.argmin(ampl)]]                                                   # choose the minimum with greatest amplitude
                        new_wvl_min_x[i] = cut_x[np.argmin(ampl)]

                    cut = (scales[:, None] == wvl_min[wvl_min_x > exact_maxima_x[-1]]).argmax(axis=0)                   # same for the last minimum
                    cut_x = wvl_min_x[(wvl_min_x > exact_maxima_x[-1])]                                                 # since it has no right minimum next to it, it is outside of the loop
                    ampl = wvl[cut,cut_x]                                                                               # if so
                    new_wvl_min[-1] = scales[cut[np.argmin(ampl)]]                                                      # choose the minimum with greatest amplitude
                    new_wvl_min_x[-1] = cut_x[np.argmin(ampl)]
                else:   
                    new_wvl_min = np.array([scales[np.where(scales==wvl_min.min())[0]]])
                    new_wvl_min_x = np.array([wvl_min_x[np.argmin(wvl_min)]])                                            # just take the global minimum in this case
                
                # calculate weight with which maxima and minima are averaged
                L = 0 
                idxs1 = np.zeros(len(exact_maxima))
                idxs2 = np.zeros(len(exact_maxima))

                idxs = np.where(np.diff(np.sign(filtered - filtered.mean())) > 0)[0]  # roots from minus to plus
                if len(idxs) > 1: 
                    wvl_maxima = exact_maxima[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]                  # only consider those extreme within the first and the last root
                    wvl_minima = new_wvl_min[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]
                    wvl_maxima_x = exact_maxima_x[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]
                    wvl_minima_x = new_wvl_min_x[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]
                    if len(wvl_maxima) > 0:
                        for g in range(len(wvl_maxima)):
                            idxs1[g] = idxs[idxs < wvl_maxima_x[g]][-1]                                                 # get all roots from minus to one 
                            idxs2[g] = idxs[idxs > wvl_minima_x[g]][0]                                                  # distance between two of thenm
                            L += idxs2[g] - idxs1[g]                                                                    # sum of all distances
                        
                        weight = (L/4 - np.sum(wvl_maxima)) / np.sum(wvl_minima - wvl_maxima)                           # weight is chosen such that the sum of all widths of the wavelet tranform is equal to the total length L
                        if weight < 0:                                                                                  # must be between 0 and 1
                            weight = .7   
                        elif weight > 1.:
                            weight = .7                                                                                 # if there is no extrema left after cutting overwrite the old extrema and choose the empirical weight = 0.7
                    
                    else:                                                                                               
                        wvl_maxima = exact_maxima
                        wvl_minima = new_wvl_min
                        wvl_maxima_x = exact_maxima_x
                        wvl_minima_x = new_wvl_min_x
                        weight = .7

                else:                                                                                            # if there is no root at all overwrite the old extrema and choose the empirical weight = 0.7
                    wvl_maxima = exact_maxima
                    wvl_minima = new_wvl_min
                    wvl_maxima_x = exact_maxima_x
                    wvl_minima_x = new_wvl_min_x
                    weight = .7


                mean_point_x = np.append(wvl_minima_x  ,wvl_maxima_x)                                        # append corrected maxima and minima
                sort_idx = np.argsort(mean_point_x)                                                           # sort along time
                mean_point_x = mean_point_x[sort_idx]    
                mean_point = np.repeat(weight*wvl_minima + (1-weight)*wvl_maxima,2)                          # replace them by the pairwise weighted average of maximum and minimum.
                widths_mean = interp1d(mean_point_x,mean_point,kind='previous',bounds_error=False, fill_value=(mean_point[0],mean_point[-1]))(np.linspace(idxs[0],idxs[-1],1000)) # linear interpolation
                freq = pywt.scale2frequency('mexh',widths_mean) / (duration/len(intensity))               # width / scale information is converted into frequencies in Hz 

                if len(freq[freq>frequency]) > 0:
                    tachycard = np.sum(freq[freq>frequency] - frequency) / len(freq)                           # average amount of tachycardic i.e higher frequency contribution
                else:
                    tachycard = 0.

                if len(freq[freq<frequency]) > 0:
                    bradycard = -np.sum(freq[freq<frequency] - frequency) / len(freq)                         # average amount of bradicardic i.e lower frequency contribution
                else:
                    bradycard = 0.

                total = np.mean(abs(freq - frequency))                                                         # total derivation from pacing frequency
                mean_freq = np.mean(freq)                                                                      # average frequency
                VarFreq = np.std(freq) * scale  / mean_freq                                                    # Variance frequency
                if VarFreq > 1.:                                                                               # to avoid to high weight on variance over orrelation & now both are on  scale of 0 to 1
                    VarFreq = 1.


                if mean_freq * duration > 2: # only possible for signals with more than 2 maxima
                    x = (filtered - filtered.mean()) / ( filtered.max() - filtered.min())                           # normalise
                    idx1 = np.where(np.diff(np.sign(x)) > 0)[0]                                                     # roots from minus to plus
                    idx2 = np.where(np.diff(np.sign(x)) < 0)[0]                                                     # roots from plus to minus
                    if (len(idx1)>0) & (len(idx2)>0):
                        x1 = x[idx1[0]:idx1[-1]]                                                                    # crop signal at first and last root from minus to plus
                        x2 = x[idx2[0]:idx2[-1]]                                                                    # crop signal at first and last root from plus to minus
                    else:
                        raise ValueError('The input signal is not as expected, since it doesn\'t cross zero periodically.') # if no root is found
                    #plt.plot(time,(intensity-intensity.mean())/(intensity.max()-intensity.min()))
                    #plt.plot(time[idx1[0]:idx1[-1]],gaussian_filter1d(x1, sigma = sigma,mode='nearest'))
                    #plt.plot(time[idx2[0]:idx2[-1]],gaussian_filter1d(x2, sigma = sigma,mode='nearest'))
                    #plt.show()    
                    autocorr1 = np.correlate(np.append(x1,x1), x1, mode='valid')                                    # perform valid autocorrelation with tiled signal
                    autocorr2 = np.correlate(np.append(x2,x2), x2, mode='valid')  

                    autocorr1 -= autocorr1.min()                                                                    # normalise
                    autocorr1 /= autocorr1.max()
                    autocorr2 -= autocorr2.min()
                    autocorr2 /= autocorr2.max()
                    maxima1,prop = find_peaks(autocorr1,prominence=prominence)                                           # find maxima
                    #corr1 = np.sum(autocorr1[maxima1]*prop['prominences']/np.sum(prop['prominences']))              # calculate weighted average of the maxima with respect to the prominences 
                    corr1 = np.mean(autocorr1[maxima1])              # calculate weighted average of the maxima with respect to the prominences 

                    maxima2,prop = find_peaks(autocorr2,prominence=prominence)
                    #corr2 = np.sum(autocorr2[maxima2]*prop['prominences']/np.sum(prop['prominences']))
                    corr2 = np.mean(autocorr2[maxima2])

                    if corr1 > corr2:                                                                               # take the correlation that resulted in a higher value
                        corr = corr1                                                                                # to reduce errors in this process
                        autocorr = autocorr1
                        maxima = maxima1
                        idx = idx1
                        cropping = '- to +'
                    else:
                        corr = corr2
                        autocorr = autocorr2  
                        maxima = maxima2   
                        idx = idx2     
                        cropping = '+ to -'

                else:
                    print('\tWarning: Transient has less than three peaks, the algorith can\'t determine correlation.\n')
                    corr = np.nan

                # invert scale
                corr = 1. - corr

                weights[c]= weight
                croppings[c] = cropping
                sigmas[c] = sigma
                reds[c] = red

                # check if any of the paameters are NaNs, if so only take the finite one
                if (np.isfinite(VarFreq)) & (np.isfinite(corr)):
                    arrythmia = varfreq_weight * VarFreq + (1 - varfreq_weight) * corr  
                elif (~np.isfinite(VarFreq)) & (np.isfinite(corr)):
                    arrythmia = corr                                                           
                elif (np.isfinite(VarFreq)) & (~np.isfinite(corr)):
                    arrythmia = VarFreq                                                           
                else:
                    arrythmia = np.nan 

                # Classification based on thresholds
                if arrythmia > threshold_arrythmia:
                    if (tachycard > threshold_freq) & (bradycard < threshold_freq):
                        classification = 'Tachycardic Arrhythmic'
                    elif (tachycard < threshold_freq) & (bradycard > threshold_freq):
                        classification = 'Bradycardic Arrhythmic'
                    else: 
                        classification = 'Arrhythmic'

                elif arrythmia <= threshold_arrythmia:
                    if (tachycard > threshold_freq) & (bradycard < threshold_freq):
                        classification = 'Tachycardic Rhythmic'
                    elif (tachycard < threshold_freq) & (bradycard > threshold_freq):
                        classification = 'Bradycardic Rhythmic'
                    else: 
                        classification = 'Rhythmic'
                else:
                    classification = 'failed'

                #print(f'\t Mean Frequency = {mean_freq} Hz')
                #print(f'\t Tachycardic Frequency deviation = {tachycard} Hz')
                #print(f'\t Bradycardic Frequency deviation = {bradycard} Hz')
                #print(f'\t Absolute Frequency deviation = {total} Hz')
                #print(f'\t Inverted Mean Correlation = {corr}')
                #print(f'\t Frequency Variance = {VarFreq}')
                #print(f'\t Arrhythmia = {arrythmia}')
                #print('\t Classification: '+classification)
                #print('\n')


                #-------------------PLOTTING---------------------------------------------------------------------------------
                if show:
                    fig,ax = plt.subplots(3,1,figsize=(12,8))
                    ax[0].plot(time,intensity-filtered.mean(),'-k')
                    ax[0].plot(time,filtered-filtered.mean(),label='Gaussian filter ($\\sigma = $' + str(sigma)+')')
                    ax[0].axvline(time[idx[0]],ls='--',color='black')
                    ax[0].axvline(time[idx[-1]],ls='--',color='black')
                    ax[0].set_ylabel('Intensity')
                    ax[0].set_xlabel('time \ s')
                    ax[0].legend(loc='upper center', bbox_to_anchor=(0.5, -0.05),fancybox=True, shadow=True, ncol=1)
                    
                    ax[1].plot(np.linspace(idxs[0],idxs[-1],1000),freq,'-')
                    ax[1].axhline(frequency,color='black')
                    ax[1].axhline(np.mean(freq),ls='--',color='red',label='mean frequency')
                    ax[1].set_ylim(0,max(freq.max(),frequency)+.2)
                    ax[1].set_xlim(0,len(intensity))
                    ax[1].set_xlabel('time \s')
                    ax[1].set_ylabel('Frequency \ Hz')
                    ax[1].legend(loc='upper center', bbox_to_anchor=(0.5, -0.05),fancybox=True, shadow=True, ncol=1)
                    if mean_freq * duration > 2:
                        ax[2].plot(autocorr)
                        ax[2].axhline(corr,ls='--',color='black')
                        ax[2].plot(maxima,autocorr[maxima],'.r')
                        ax[2].set_xlabel('lag time')
                        ax[2].set_ylabel('Correlation')
                    plt.show()
                    plt.cla()
                    plt.close('all')
                
                if not os.path.exists(path_figures):
                    os.makedirs(path_figures)


                fig,ax = plt.subplots(1,3,figsize=(16,2))

                ax[0].plot(time,intensity-filtered.mean(),'-k')
                ax[0].plot(time,filtered-filtered.mean(),label='Gaussian filter ($\\sigma = $' + str(sigma)+')')
                ax[0].axvline(time[idx[0]],ls='--',color='black')
                ax[0].axvline(time[idx[-1]],ls='--',color='black')
                ax[0].set_ylabel('Intensity')
                ax[0].set_xlabel('time \ s')
                
                ax[1].plot(np.linspace(idxs[0],idxs[-1],1000),freq,'-')
                ax[1].axhline(frequency,color='black')
                ax[1].axhline(np.mean(freq),ls='--',color='red',label='mean frequency')
                ax[1].set_ylim(0,max(freq.max(),frequency)+.2)
                ax[1].set_xlim(0,len(intensity))
                ax[1].set_xlabel('time \s')
                ax[1].set_ylabel('Frequency \ Hz')
                if mean_freq * duration > 2:
                    ax[2].plot(autocorr)
                    ax[2].axhline(1-corr,ls='--',color='black')
                    ax[2].plot(maxima,autocorr[maxima],'.r')
                    ax[2].set_xlabel('lag time')
                    ax[2].set_ylabel('Correlation')
                plt.tight_layout()
                plt.savefig(os.path.join(path_figures,os.path.basename(file)[:-4]+'.png'),dpi=300)
                plt.cla()
                plt.close('all')
                #-------------------PLOTTING---------------------------------------------------------------------------------
                im = openpyxl.drawing.image.Image(os.path.join(path_figures,os.path.basename(file)[:-4]+'.png')) # include the plot
                im.height= 80
                im.width = 1000
                sheet.add_image(im,"N"+str(position+1))

                sheet.cell(row=1,column=3).value = 'absolute deviation'
                sheet.cell(row=1,column=4).value = 'Mean frequency'
                sheet.cell(row=1,column=5).value = 'Frequeny Variance'
                sheet.cell(row=1,column=6).value = 'Bradycardia'
                sheet.cell(row=1,column=7).value = 'Tachycardia'
                sheet.cell(row=1,column=8).value = 'Correlation'
                sheet.cell(row=1,column=9).value = 'Arrhythmia'
                sheet.cell(row=1,column=10).value = 'Classification'

                sheet.cell(row=c+3,column=1).value =f'File {c+1}: '+ file.split('\\')[-1]
                sheet.cell(row=position+1,column=13).value =f'File {c+1}'

                sheet.cell(row=c+3,column=3).value = np.round(total,5)
                sheet.cell(row=c+3,column=4).value = np.round(mean_freq,5)
                sheet.cell(row=c+3,column=5).value = np.round(VarFreq,5)
                sheet.cell(row=c+3,column=6).value = np.round(bradycard,5)
                sheet.cell(row=c+3,column=7).value = np.round(tachycard,5)
                sheet.cell(row=c+3,column=8).value = np.round(corr,5)
                sheet.cell(row=c+3,column=9).value = np.round(arrythmia,5)
                sheet.cell(row=c+3,column=10).value = classification

                position += 5
                histogram[c,0] = total
                histogram[c,1] = arrythmia
                distribution[classes[classification]] += 1
            except:
                if save_mode:
                    raise ValueError("Something went wrong with the file! Check if your file is not corrupted.")
                else:
                    print('Something went wrong with this file, I\'m skipping it!')
                    continue
        #histogram = histogram[:c]
        plt.hist(histogram[:,0],bins=20,color='royalblue', alpha=0.7)
        plt.axvline(threshold_freq,color='black',ls='--')
        plt.xlim(left=0,right=1.1)
        plt.xlabel('Total Frequency Deviation', fontsize=14, fontweight='bold')
        plt.ylabel('Frequency',fontsize=14, fontweight='bold')
        sns.despine()
        plt.savefig(os.path.join(path_figures,'hist1.png'),dpi=300)
        plt.cla()
        plt.close('all')
        img = openpyxl.drawing.image.Image(os.path.join(path_figures,'hist1.png')) # include the plot
        img.height= 200
        img.width = 250
        sheet.add_image(img,"B"+str(c+5))

        plt.hist(histogram[:,1],bins=20,color='royalblue', alpha=0.7)
        plt.axvline(threshold_arrythmia,color='black',ls='--')
        plt.xlim(left=0,right=1.1)
        plt.xlabel('Arrythmia [A.U.]',fontsize=14, fontweight='bold')
        plt.ylabel('Frequency',fontsize=14, fontweight='bold')
        sns.despine()
        plt.savefig(os.path.join(path_figures,'hist2.png'),dpi=300)
        plt.cla()
        plt.close('all')
        img = openpyxl.drawing.image.Image(os.path.join(path_figures,'hist2.png')) # include the plot
        img.height= 200
        img.width = 250
        sheet.add_image(img,"B"+str(c+18))

        colors = ['seagreen'] + ['indianred']*2 + ['maroon']*3
        sns.barplot(x= np.arange(len(distribution)), y=distribution, palette=colors)
        x_labels = ['Rhythmic' , 'Tachycardic\nRhythmic', 'Bradycardic\nRhythmic', 'Arrhythmic', 'Tachycardic\nArrhythmic', 'Bradycardic\nArrhythmic']
        plt.ylabel("Frequency", fontsize=14, fontweight='bold')
        plt.xticks(np.arange(len(distribution)), x_labels, rotation=45, fontsize=10, fontweight='bold')
        sns.despine()
        plt.tight_layout()
        plt.savefig(path_figures+'\\hist3.png',dpi=300)
        plt.cla()
        plt.close('all')
        img = openpyxl.drawing.image.Image(path_figures+'\\hist3.png') # include the plot
        img.height= 200
        img.width = 250
        sheet.add_image(img,"G"+str(c+6))        
        
        print('\nSaving the output files to '+path+' ...\n')
        workbook.save(filename=path+'\\Arrhythpy_analysis.xlsx')      

    #os.remove(path+'\\hist.png')
    else:
        raise ValueError('Sorry, I did not find any files.')

def is_numeric(s):
    s = s.lstrip('-+')
    return (s.isdigit() and '.' not in s) or (s.count('.') == 1 and all(part.isdigit() for part in s.split('.')))


def load_transient(file):
    try:
        data_type = file.split('.')[-1]
        if data_type =='lsm' or data_type == 'tif':
            img = imread(file)[:,0,:]
            intensity = np.mean(img,axis=1)
        elif data_type == 'csv':
            intensity = np.genfromtxt(file, delimiter=' ')

        red = int(len(intensity) // 5000)                                                    # this is to reduce computation time
        if red < 1:
            red = 1

        intensity = intensity[::red]
        return intensity
    except:
        return np.nan

def analyse_single(intensity,frequency, duration = 10.,sigma0 = .05,prominence=0.0005,threshold_abs = .1,varfreq_weight = 0.5,scale = 1.):
    '''
    Parameter:
    - file:                 single file to be analysed. Must be line scan images (either tif or lsm files)
    - frequency:   	        float; pacing frequency or known eigenfrequency of cells
    - duration:             float; duration of the transient in signal (needed for unit conversion)
    - sigma0:  	            float; Standard deviation of the Gaussian used for the gaussian filter. Too low values will increase the noise in the signal,while too large values will eliminate highly tachycardic signals. 
                                sigma is changed depending on the largest amplitude of the Fourier specturm to adapt it to the acutal frequency in the signal.
    - threshold_abs:        float; threshold for local extrema i.e. minimum height of the peak
    - varfreq_weight:       float; weight in weithed average of VF and IMC
    - scale:                float; the VF is scaled by this number to have IMC and VF roughly on th same scale.
    '''

    try:
        fft = abs(np.fft.rfft(intensity-intensity.mean())) 
        freq_fft = np.fft.fftfreq(len(intensity), d=duration/len(intensity))[np.argmax(fft)] / 2        # get maximum amplitude of fourier spectrum , /2 since rfft crops the symetric spectrum at half

        if frequency is None:
            frequency = freq_fft
            print('Warning: No pacing/eigen frequency given. Tachycardia and bradycardia are calculated with highest peak in fourier spectrum. \n')

        width_used = len(intensity)/(duration*frequency) / 4                                        # for mexical hat wavelet width is defined like this
        #end_width = max(width_used+50,len(intensity)/(duration*freq_fft)/3)  # largest width used for wavelet tranform (depending on ratio between given frequency and frequency of main fourier coefficient)
        end_width = len(intensity)/(duration*freq_fft) / 3  # largest width used for wavelet tranform (depending on ratio between given frequency and frequency of main fourier coefficient)

        sigma = int (len(intensity)/(duration*freq_fft) *sigma0)             # width of gaussian is adapted to main frequency component
        filtered = gaussian_filter1d(intensity,sigma=sigma,mode='nearest')   # gaussian filter to smooth

        scales = np.linspace(1,end_width,500)
        wvl,_ = pywt.cwt(intensity-np.mean(intensity),scales,'mexh',sampling_period=duration/len(intensity))  # continious wavelet transform
        wvl /= (wvl.max() - wvl.min())     
        wvl = gaussian_filter(wvl,10)     

        # 1.
        mins = peak_local_max(-wvl,min_distance=10,threshold_abs=threshold_abs,exclude_border=False)    # locate local minima
        wvl_min_x,wvl_min = mins[:,1],scales[mins[:,0]] # cut off the sides to avoid border artifacts
        wvl_min = wvl_min[np.argsort(wvl_min_x)]                                                        # sort along time
        wvl_min_x = wvl_min_x[np.argsort(wvl_min_x)]
        wvl_min = wvl_min[wvl_min_x<len(intensity)-width_used*2]
        wvl_min_x = wvl_min_x[wvl_min_x<len(intensity)-width_used*2]

        maxs = peak_local_max(wvl,min_distance=10,threshold_abs=threshold_abs ,exclude_border=False)    # locate local maxima
        wvl_max_x,wvl_max = maxs[:,1],scales[maxs[:,0]] # cut off the sides to avoid border artifacts
        wvl_max = wvl_max[np.argsort(wvl_max_x)]                                                        # sort along time
        wvl_max_x = wvl_max_x[np.argsort(wvl_max_x)]
        wvl_max = wvl_max[wvl_max_x>width_used*2]
        wvl_max_x = wvl_max_x[wvl_max_x>width_used*2]

        # 2.
        if len(wvl_min) > 1: 
            f_diff = interp1d(wvl_min_x,wvl_min)                                                        # linear interpolation
            wvl_min = wvl_min[wvl_min_x>= wvl_max_x[0]]                                                 # delete minima left to first maximum
            wvl_min_x = wvl_min_x[wvl_min_x>= wvl_max_x[0]]                                             # since we have to start with a maximum
            exact_maxima = np.zeros(wvl_min_x.shape)
            exact_maxima_x = np.zeros(wvl_min_x.shape)

            for i,_ in enumerate(wvl_min_x):
                if i == 0:                                                                              # for the first minimum we don't have a minimum before that -> just use the average distance
                    approx_max_x = wvl_min_x[i] - .5*np.mean(np.diff(wvl_min_x))
                    approx_max = wvl_min[i]
                else:
                    approx_max_x = wvl_min_x[i] - .5 * np.diff([wvl_min_x[i-1],wvl_min_x[i]])[0]        # get middle point between two minima for obtain an approximate maximum
                    approx_max = f_diff(approx_max_x)  

                dist = (wvl_max_x[wvl_max_x < wvl_min_x[i]] - approx_max_x)**2 + (wvl_max[wvl_max_x < wvl_min_x[i]] - approx_max)**2 # get dsitance betweeen the approximate maxima and the actual maxima
                exact_maxima[i] = wvl_max[np.argmin(dist)]                                              # choose the maximum closest to the approximate one
                exact_maxima_x[i] = wvl_max_x[np.argmin(dist)]
        else:                                                                                           # if only one entery, then I don't need this and I just take the global maximum
            exact_maxima,exact_maxima_x = np.where(wvl==wvl.max())
        
        exact_maxima_x,idx_unique = np.unique(exact_maxima_x,return_index=True)                         # reduce the array to only contain unique enteries
        exact_maxima = exact_maxima[idx_unique]

        # 3.
        for i in range(len(wvl_min)-1):
            cut = exact_maxima[(exact_maxima_x > wvl_min_x[i]) & (exact_maxima_x < wvl_min_x[i+1])]     # cut between two minima 
            cut_x = exact_maxima_x[(exact_maxima_x > wvl_min_x[i]) & (exact_maxima_x < wvl_min_x[i+1])] # to see if there is more than one minimum
            if len(cut) >1:                                                                             # if so 
                d = (cut_x - wvl_min_x[i+1])**2 + (cut - wvl_min[i+1])**2                               # calculate the distance                     
                exact_maxima = np.delete(exact_maxima,(exact_maxima_x[:, None] == cut_x[cut_x != cut_x[np.argmin(d)]]).argmax(axis=0)) # delete those maxima, that are not the closest one
                exact_maxima_x = np.delete(exact_maxima_x,(exact_maxima_x[:, None] == cut_x[cut_x != cut_x[np.argmin(d)]]).argmax(axis=0))

        # 4.
        new_wvl_min = np.zeros(exact_maxima.shape)
        new_wvl_min_x = np.zeros(exact_maxima.shape)
        if len(exact_maxima) > 1:
            for i in range(len(exact_maxima)-1):

                cut = (scales[:, None] == wvl_min[(wvl_min_x > exact_maxima_x[i]) & (wvl_min_x < exact_maxima_x[i+1])]).argmax(axis=0) # float scale values are converted into indeces
                cut_x = wvl_min_x[(wvl_min_x > exact_maxima_x[i]) & (wvl_min_x < exact_maxima_x[i+1])]          # to see if there ise more than one minimum between two maxima
                ampl = wvl[cut,cut_x]                                                                           # if so
                new_wvl_min[i] = scales[cut[np.argmin(ampl)]]                                                   # choose the minimum with greatest amplitude
                new_wvl_min_x[i] = cut_x[np.argmin(ampl)]

            cut = (scales[:, None] == wvl_min[wvl_min_x > exact_maxima_x[-1]]).argmax(axis=0)                   # same for the last minimum
            cut_x = wvl_min_x[(wvl_min_x > exact_maxima_x[-1])]                                                 # since it has no right minimum next to it, it is outside of the loop
            ampl = wvl[cut,cut_x]                                                                               # if so
            new_wvl_min[-1] = scales[cut[np.argmin(ampl)]]                                                      # choose the minimum with greatest amplitude
            new_wvl_min_x[-1] = cut_x[np.argmin(ampl)]
        else:   
            new_wvl_min = np.array([scales[np.where(scales==wvl_min.min())[0]]])
            new_wvl_min_x = np.array([wvl_min_x[np.argmin(wvl_min)]])                                            # just take the global minimum in this case
        
        # calculate weight with which maxima and minima are averaged
        L = 0 
        idxs1 = np.zeros(len(exact_maxima))
        idxs2 = np.zeros(len(exact_maxima))

        idxs = np.where(np.diff(np.sign(filtered - filtered.mean())) > 0)[0]  # roots from minus to plus
        if len(idxs) > 1: 
            wvl_maxima = exact_maxima[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]                  # only consider those extreme within the first and the last root
            wvl_minima = new_wvl_min[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]
            wvl_maxima_x = exact_maxima_x[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]
            wvl_minima_x = new_wvl_min_x[(exact_maxima_x > idxs[0]) & (new_wvl_min_x < idxs[-1])]
            if len(wvl_maxima) > 0:
                for g in range(len(wvl_maxima)):
                    idxs1[g] = idxs[idxs < wvl_maxima_x[g]][-1]                                                 # get all roots from minus to one 
                    idxs2[g] = idxs[idxs > wvl_minima_x[g]][0]                                                  # distance between two of thenm
                    L += idxs2[g] - idxs1[g]                                                                    # sum of all distances
                
                weight = (L/4 - np.sum(wvl_maxima)) / np.sum(wvl_minima - wvl_maxima)                           # weight is chosen such that the sum of all widths of the wavelet tranform is equal to the total length L
                if weight < 0:                                                                                  # must be between 0 and 1
                    weight = .7   
                elif weight > 1.:
                    weight = .7                                                                                 # if there is no extrema left after cutting overwrite the old extrema and choose the empirical weight = 0.7
            
            else:                                                                                               
                wvl_maxima = exact_maxima
                wvl_minima = new_wvl_min
                wvl_maxima_x = exact_maxima_x
                wvl_minima_x = new_wvl_min_x
                weight = .7

        else:                                                                                            # if there is no root at all overwrite the old extrema and choose the empirical weight = 0.7
            wvl_maxima = exact_maxima
            wvl_minima = new_wvl_min
            wvl_maxima_x = exact_maxima_x
            wvl_minima_x = new_wvl_min_x
            weight = .7


        mean_point_x = np.append(wvl_minima_x  ,wvl_maxima_x)                                        # append corrected maxima and minima
        sort_idx = np.argsort(mean_point_x)                                                           # sort along time
        mean_point_x = mean_point_x[sort_idx]    
        mean_point = np.repeat(weight*wvl_minima + (1-weight)*wvl_maxima,2)                          # replace them by the pairwise weighted average of maximum and minimum.
        widths_mean = interp1d(mean_point_x,mean_point,kind='previous',bounds_error=False, fill_value=(mean_point[0],mean_point[-1]))(np.linspace(idxs[0],idxs[-1],1000)) # linear interpolation
        freq = pywt.scale2frequency('mexh',widths_mean) / (duration/len(intensity))               # width / scale information is converted into frequencies in Hz 

        if len(freq[freq>frequency]) > 0:
            tachycard = np.sum(freq[freq>frequency] - frequency) / len(freq)                           # average amount of tachycardic i.e higher frequency contribution
        else:
            tachycard = 0.

        if len(freq[freq<frequency]) > 0:
            bradycard = -np.sum(freq[freq<frequency] - frequency) / len(freq)                         # average amount of bradicardic i.e lower frequency contribution
        else:
            bradycard = 0.

        mean_freq = np.mean(freq)                                                                      # average frequency
        VarFreq = np.std(freq) * scale  / mean_freq                                                    # Variance frequency
        if VarFreq > 1.:                                                                               # to avoid to high weight on variance over orrelation & now both are on  scale of 0 to 1
            VarFreq = 1.


        if mean_freq * duration > 2: # only possible for signals with more than 2 maxima
            x = (filtered - filtered.mean()) / ( filtered.max() - filtered.min())                           # normalise
            idx1 = np.where(np.diff(np.sign(x)) > 0)[0]                                                     # roots from minus to plus
            idx2 = np.where(np.diff(np.sign(x)) < 0)[0]                                                     # roots from plus to minus
            if (len(idx1)>0) & (len(idx2)>0):
                x1 = x[idx1[0]:idx1[-1]]                                                                    # crop signal at first and last root from minus to plus
                x2 = x[idx2[0]:idx2[-1]]                                                                    # crop signal at first and last root from plus to minus
            else:
                raise ValueError('The input signal is not as expected, since it doesn\'t cross zero periodically.') # if no root is found
            #plt.plot(time,(intensity-intensity.mean())/(intensity.max()-intensity.min()))
            #plt.plot(time[idx1[0]:idx1[-1]],gaussian_filter1d(x1, sigma = sigma,mode='nearest'))
            #plt.plot(time[idx2[0]:idx2[-1]],gaussian_filter1d(x2, sigma = sigma,mode='nearest'))
            #plt.show()    
            autocorr1 = np.correlate(np.append(x1,x1), x1, mode='valid')                                    # perform valid autocorrelation with tiled signal
            autocorr2 = np.correlate(np.append(x2,x2), x2, mode='valid')  

            autocorr1 -= autocorr1.min()                                                                    # normalise
            autocorr1 /= autocorr1.max()
            autocorr2 -= autocorr2.min()
            autocorr2 /= autocorr2.max()
            maxima1,prop = find_peaks(autocorr1,prominence=prominence)                                           # find maxima
            #corr1 = np.sum(autocorr1[maxima1]*prop['prominences']/np.sum(prop['prominences']))              # calculate weighted average of the maxima with respect to the prominences 
            corr1 = np.mean(autocorr1[maxima1])              # calculate weighted average of the maxima with respect to the prominences 

            maxima2,prop = find_peaks(autocorr2,prominence=prominence)
            #corr2 = np.sum(autocorr2[maxima2]*prop['prominences']/np.sum(prop['prominences']))
            corr2 = np.mean(autocorr2[maxima2])

            if corr1 > corr2:                                                                               # take the correlation that resulted in a higher value
                corr = corr1                                                                                # to reduce errors in this process
            else:
                corr = corr2
        else:
            print('\tWarning: Transient has less than three peaks, the algorith can\'t determine correlation.\n')
            corr = np.nan

        # invert scale
        corr = 1. - corr

        # check if any of the paameters are NaNs, if so only take the finite one
        if (np.isfinite(VarFreq)) & (np.isfinite(corr)):
            arrythmia = varfreq_weight * VarFreq + (1 - varfreq_weight) * corr  
        elif (~np.isfinite(VarFreq)) & (np.isfinite(corr)):
            arrythmia = corr                                                           
        elif (np.isfinite(VarFreq)) & (~np.isfinite(corr)):
            arrythmia = VarFreq                                                           
        else:
            arrythmia = np.nan 
        return arrythmia, tachycard, bradycard
    except:
        return np.nan, np.nan, np.nan
    
def get_threshold_freq(classification,tachycards,bradycards,plot=False):
    '''
    performs brute force algorithm to find optimal threshold, where all deviations (tachycard and bradycard) are lumped together into one value
    '''
    class_freq_human = np.zeros(classification.shape)
    class_freq_human[(classification==2.) | (classification==3.) | (classification==5.) | (classification==6.)] = 1.
    
    dev = np.dstack((tachycards,bradycards))[0] # lump bradycard and tachycard contribution together
    freq_measure = np.max(dev,axis=1)           # into one frequency measure

    threshs = np.linspace(0,freq_measure.max(),1000)
    accuracies = np.zeros(threshs.shape)
    for i,thresh in enumerate(threshs):
        class_temp = np.zeros(class_freq_human.shape)
        class_temp[freq_measure>thresh] = 1.
        accuracies[i] = accuracy_score(class_freq_human,class_temp)

    max_indices = np.where(accuracies == np.max(accuracies))[0]
    middle_index = max_indices[len(max_indices) // 2]
    thresh_opt_freq = threshs[middle_index]
    class_final = np.zeros(class_freq_human.shape)
    class_final[freq_measure>thresh_opt_freq] = 1.
    if plot:
        plt.plot(freq_measure,class_freq_human,'.')
        plt.plot(threshs,accuracies,'-')
        plt.axvline(thresh_opt_freq)
        plt.show()
    return thresh_opt_freq,accuracy_score(class_freq_human,class_final)

# make it a logisitc regressiondirectlyy using 6 labels and 2 varialbes
def get_threshold_arrythmia(classification,arrythmias,plot=False):
    '''
    performs brute force algorithm to find optimal threshold
    '''
    class_arrythmia_human = np.zeros(classification.shape)
    class_arrythmia_human[classification>=3.] = 1.
    
    threshs = np.linspace(0,arrythmias.max(),1000)
    accuracies = np.zeros(threshs.shape)
    for i,thresh in enumerate(threshs):
        class_temp = np.zeros(class_arrythmia_human.shape)
        class_temp[arrythmias>thresh] = 1.
        accuracies[i] = accuracy_score(class_arrythmia_human,class_temp)

    max_indices = np.where(accuracies == np.max(accuracies))[0]
    middle_index = max_indices[len(max_indices) // 2]
    thresh_opt_arr = threshs[middle_index]
    class_final = np.zeros(class_arrythmia_human.shape)
    class_final[arrythmias>thresh_opt_arr] = 1.
    if plot:
        plt.plot(arrythmias,class_arrythmia_human,'.')
        plt.plot(threshs,accuracies,'-')
        plt.axvline(thresh_opt_arr)
        plt.show()
    return thresh_opt_arr,accuracy_score(class_arrythmia_human,class_final)

def calc_accuracy(arrythmias,tachycardia,bradycardia, classification_human,threshold_arrhythmia,threshold_freq):
    mask_arrythmia = arrythmias > threshold_arrhythmia
    mask_tachy = tachycardia > threshold_freq
    mask_brady = bradycardia > threshold_freq
    classification = np.ones(arrythmias.shape)
    classification[mask_arrythmia & mask_tachy & ~mask_brady] = 5
    classification[mask_arrythmia & ~mask_tachy & mask_brady] = 6
    classification[mask_arrythmia & (~(mask_tachy | mask_brady) | (mask_tachy & mask_brady))] = 4 # arrhythmic but no deviation or deviations in both brady & tachy
    classification[~mask_arrythmia & mask_tachy & ~mask_brady] = 2
    classification[~mask_arrythmia & ~mask_tachy & mask_brady] = 3
    return accuracy_score(classification_human, classification)


if __name__ == "__main__":
    #------------PARAMETERS------------------------------------------------------------------
    print('\n')
    path = input('Path:\t')
    print('\n')

    file_path = os.path.join(path, "config.yaml")
    if os.path.exists(file_path):
        print("Parameter file found in folder: "+os.path.basename(path))
        with open(os.path.join(path,"config.yaml"), 'r') as yaml_file:
            params = yaml.safe_load(yaml_file)
            frequency = params["frequency"]
            duration = params["duration"]
            threshold_freq = params["threshold_freq"]
            threshold_arrythmia = params["threshold_arrythmia"]

            show = params["show"]
            sigma0 = params["sigma0"]
            threshold_abs = params["threshold_abs"]
            prominence = params["prominence"]
            varfreq_weight = params["varfreq_weight"]
            scale = params["scale"]
    else:
        print("No paramter file found. Specify the following paramters:\n")
        frequency = input('Pacing / Eigen frequency:\t')
        if is_numeric(frequency):
            frequency = float(frequency)
        print('\n')
        duration = float(input('Duration of signal in sec:\t'))
        print('\n')
        threshold_freq = float(input('Threshold for Tachycardia / Bradicardia in Hz:\t'))
        print('\n')
        threshold_arrythmia = float(input('Threshold for Arrythmia:\t'))
        print('\n')

        show = False  
        sigma0 = .01    
        threshold_abs = .1  
        prominence = 0.05 
        varfreq_weight = 0.5    
        scale = 1.  

        params = {
            'frequency': frequency,
            'duration': duration,
            'threshold_freq': threshold_freq,
            'threshold_arrythmia': threshold_arrythmia,
            'show': show,               	        # set to True if you want to check the plots and edit them.
            'sigma0': sigma0,                       # inital width of gaussian smoothing
            'threshold_abs': threshold_abs,         # threshold for peak detection of wavelet transform
            'prominence': prominence,               # omit any peaks in the correlation smaller than this value
            'varfreq_weight': varfreq_weight,       # weight of VF in weight sum of IMC and VF
            'scale': scale                          # scales the VF distribution before cropping to e [0,1]
        }
        with open(os.path.join(path,'config.yaml'), 'w') as yaml_file:
            yaml.dump(params, yaml_file)

    #------------PARAMETERS------------------------------------------------------------------

    files = glob.glob(path+'\*')
    run(files, frequency = frequency, threshold_freq = threshold_freq, threshold_arrythmia = threshold_arrythmia,
         duration = duration, sigma0 = sigma0,prominence = prominence, threshold_abs = threshold_abs, varfreq_weight = varfreq_weight, scale = scale, show = show)
